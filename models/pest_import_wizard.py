import base64
import io
from datetime import datetime

from odoo import models, fields, api
from odoo.exceptions import UserError

try:
    import openpyxl
    import openpyxl.styles
except ImportError:
    openpyxl = None


class PestIncidentImportWizard(models.TransientModel):
    _name = 'pest.incident.import.wizard'
    _description = 'Wizard para importar incidencias desde Excel'

    blueprint_id = fields.Many2one('pest.blueprint', string='Plano', required=True)
    excel_file = fields.Binary(string='Archivo Excel')
    excel_filename = fields.Char(string='Nombre del archivo')
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('preview', 'Vista Previa'),
        ('done', 'Completado'),
    ], default='draft', string='Estado')
    preview_line_ids = fields.One2many(
        'pest.incident.import.preview', 'wizard_id', string='Vista Previa')
    error_count = fields.Integer(compute='_compute_counts', string='Errores')
    success_count = fields.Integer(compute='_compute_counts', string='Válidos')
    total_count = fields.Integer(compute='_compute_counts', string='Total')

    @api.depends('preview_line_ids', 'preview_line_ids.is_valid')
    def _compute_counts(self):
        for rec in self:
            lines = rec.preview_line_ids
            rec.total_count = len(lines)
            rec.success_count = len(lines.filtered('is_valid'))
            rec.error_count = rec.total_count - rec.success_count

    def action_download_template(self):
        """Generate and download an XLSX template for this blueprint."""
        self.ensure_one()
        if not openpyxl:
            raise UserError('La librería openpyxl no está instalada.')

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        traps = self.blueprint_id.trap_ids.filtered('active')
        # Group traps by type
        trap_types = {}
        for trap in traps:
            type_name = trap.trap_type_id.name if trap.trap_type_id else 'Sin Tipo'
            if type_name not in trap_types:
                trap_types[type_name] = []
            trap_types[type_name].append(trap)

        if not trap_types:
            trap_types['General'] = []

        for type_name, type_traps in trap_types.items():
            # Sheet name max 31 chars
            sheet_name = type_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            # Headers
            headers = [
                'Trampa',
                'Fecha (YYYY-MM-DD)',
                'Tipo Plaga',
                'Tipo Incidencia (captura/hallazgo)',
                'Cantidad',
                'Notas',
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            # Pre-fill trap names
            for row, trap in enumerate(type_traps, 2):
                ws.cell(row=row, column=1, value=trap.name)
            # Auto-width
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create attachment for download
        attachment = self.env['ir.attachment'].create({
            'name': f'plantilla_incidencias_{self.blueprint_id.name}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_upload(self):
        """Parse the uploaded Excel file and create preview lines."""
        self.ensure_one()
        if not openpyxl:
            raise UserError('La librería openpyxl no está instalada.')
        if not self.excel_file:
            raise UserError('Por favor seleccione un archivo Excel.')

        # Clear previous preview
        self.preview_line_ids.unlink()

        file_data = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)

        # Build lookup maps
        trap_map = {
            t.name.strip().lower(): t
            for t in self.blueprint_id.trap_ids.filtered('active')
        }
        plague_map = {
            p.name.strip().lower(): p
            for p in self.env['pest.plague.type'].search([('active', '=', True)])
        }

        lines_vals = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=False):
                values = [cell.value for cell in row]
                # Skip empty rows
                if not any(values):
                    continue

                trap_name = str(values[0] or '').strip()
                date_raw = values[1] if len(values) > 1 else None
                plague_name = str(values[2] or '').strip() if len(values) > 2 else ''
                incident_type_raw = str(values[3] or '').strip().lower() if len(values) > 3 else ''
                organism_count_raw = values[4] if len(values) > 4 else None
                notes = str(values[5] or '').strip() if len(values) > 5 else ''

                # Match trap
                trap = trap_map.get(trap_name.lower())
                error_msg = ''

                if not trap_name:
                    error_msg = 'Nombre de trampa vacío.'
                elif not trap:
                    error_msg = f'Trampa "{trap_name}" no encontrada en este plano.'

                # Match plague type
                plague = plague_map.get(plague_name.lower())
                if not plague and not error_msg:
                    error_msg = f'Tipo de plaga "{plague_name}" no encontrado.'

                # Parse date
                parsed_date = False
                if isinstance(date_raw, datetime):
                    parsed_date = date_raw.date()
                elif hasattr(date_raw, 'year'):
                    # date object
                    parsed_date = date_raw
                elif isinstance(date_raw, str):
                    try:
                        parsed_date = datetime.strptime(date_raw.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        if not error_msg:
                            error_msg = f'Fecha inválida: "{date_raw}". Use formato YYYY-MM-DD.'
                elif date_raw is None:
                    if not error_msg:
                        error_msg = 'Fecha vacía.'

                # Validate incident type
                incident_type = incident_type_raw if incident_type_raw in ('captura', 'hallazgo') else False
                if not incident_type and not error_msg:
                    error_msg = (
                        f'Tipo de incidencia inválido: "{incident_type_raw}". '
                        'Use "captura" o "hallazgo".'
                    )

                # Validate count
                try:
                    count = int(organism_count_raw or 0)
                    if count < 0:
                        raise ValueError
                except (ValueError, TypeError):
                    count = 0
                    if not error_msg:
                        error_msg = f'Cantidad inválida: "{organism_count_raw}".'

                lines_vals.append({
                    'wizard_id': self.id,
                    'trap_name': trap_name,
                    'trap_id': trap.id if trap else False,
                    'date': parsed_date or False,
                    'plague_type_name': plague_name,
                    'plague_type_id': plague.id if plague else False,
                    'incident_type': incident_type or False,
                    'organism_count': count,
                    'notes': notes,
                    'error_message': error_msg,
                })

        if lines_vals:
            self.env['pest.incident.import.preview'].create(lines_vals)

        self.state = 'preview'
        return self._reopen_wizard()

    def action_confirm(self):
        """Create pest.incident records from valid preview lines."""
        self.ensure_one()
        valid_lines = self.preview_line_ids.filtered('is_valid')
        if not valid_lines:
            raise UserError('No hay líneas válidas para importar.')

        incident_vals = []
        for line in valid_lines:
            # Convert date to datetime for the Datetime field
            date_val = fields.Datetime.now()
            if line.date:
                date_val = datetime.combine(line.date, datetime.min.time())

            incident_vals.append({
                'trap_id': line.trap_id.id,
                'blueprint_id': self.blueprint_id.id,
                'sede_id': self.blueprint_id.sede_id.id,
                'plague_type_id': line.plague_type_id.id,
                'incident_type': line.incident_type,
                'organism_count': line.organism_count,
                'date': date_val,
                'notes': line.notes or '',
            })

        self.env['pest.incident'].create(incident_vals)
        self.state = 'done'
        return self._reopen_wizard()

    def _reopen_wizard(self):
        """Return action to reopen this wizard (keep same record)."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class PestIncidentImportPreview(models.TransientModel):
    _name = 'pest.incident.import.preview'
    _description = 'Línea de vista previa de importación'

    wizard_id = fields.Many2one('pest.incident.import.wizard', ondelete='cascade')
    trap_name = fields.Char(string='Trampa (Excel)')
    trap_id = fields.Many2one('pest.trap', string='Trampa')
    date = fields.Date(string='Fecha')
    plague_type_name = fields.Char(string='Tipo Plaga (Excel)')
    plague_type_id = fields.Many2one('pest.plague.type', string='Tipo Plaga')
    incident_type = fields.Selection([
        ('captura', 'Captura'),
        ('hallazgo', 'Hallazgo'),
    ], string='Tipo Incidencia')
    organism_count = fields.Integer(string='Cantidad')
    notes = fields.Text(string='Notas')
    error_message = fields.Char(string='Error')
    is_valid = fields.Boolean(compute='_compute_is_valid', store=True, string='Válido')

    @api.depends('trap_id', 'plague_type_id', 'incident_type', 'error_message')
    def _compute_is_valid(self):
        for line in self:
            line.is_valid = bool(
                line.trap_id and line.plague_type_id
                and line.incident_type and not line.error_message
            )
